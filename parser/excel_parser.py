import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

wb=load_workbook('template.xlsx')

dataset={
    "Батарея":{
        "123456A":[1,2,3,4,5,6],
        "123456B":[1,2,3,4,5,7]
        },
    "Микроконтроллер":{
        "123456AС":[1,2,3,4,5,6,7,8,9,10],
        "123456BВ":[1,2,3,4,5,7,8,10,11,12]
        }
}

#ws активный лист
def get_range_active_column(ws):
    column_list = [cell.column for cell in ws[1]]
    return column_list[0],len(column_list)

def parse_and_write_excel(dataset):
    global wb
    for key_component in dataset.keys():
        ws=wb[key_component]
        start_index,end_index=get_range_active_column(ws)
        item=dataset[key_component]
        current_row=2
        #цикл по URL
        for key_url in item.keys():
            data=item[key_url]
            url_cell=ws.cell(row = current_row, column = end_index-1)
            url_cell.value=key_url
            print(key_url,data,start_index,end_index)
            #индекс элемента списка данных
            index=0
            #цикл по строкам
            for row in ws.iter_rows(min_row=current_row, max_col=end_index, max_row=current_row):
                index=0
                for cell in row:
                    if (index!=end_index-2):
                        print(cell)
                        cell.value=data[index]
                    index+=1
                    
            current_row+=1       

    wb.save('result.xlsx')
            
    


if __name__ == "__main__":
    wb = load_workbook('template.xlsx')
    print(wb.sheetnames)
    parse_and_write_excel(dataset)
    # for sheet in wb.sheetnames:
    #     ws=wb[sheet]
    #     print(sheet)
    #     print("_______________________________________________")
    #     for col in ws.iter_cols(min_row=1, max_col=22, max_row=1):
    #         for cell in col:
    #             print(cell.value)
    #     # for col in ws.iter_cols(min_row=2, max_col=3, max_row=3):
    #     #     for cell in col:
    #     #         cell.value="1"
    #     #         print(cell)
    # #wb2.save('template2.xlsx')
    #         #cell.value=1
