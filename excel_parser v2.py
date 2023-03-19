import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Font
from enum import Enum, auto

class CompEnum(Enum):
    """Перечисления компонентов"""

    BATTERY = auto()
    MICROCONTROLLER = auto()
    ELECTRICMOTOR = auto()
    MOTORCONTROLLER = auto()
    FLIGHTCONTROLLER = auto()
    LIDAR = auto()
    MICROFLIGHTCONTROLLER = auto()
    RANGEFINDER = auto()
    SATELLITECOMMMODULE = auto()
    LEASHINGPLATFORM = auto()
    THERMALCAMERA = auto()
    UAVCOPTERTYPE = auto()
    VIDEOTRANSMITTER = auto()
    PAYLOAD = auto()
    CONTROLPANEL = auto()

wb=load_workbook('template.xlsx')

correspondence_dict={
    CompEnum.BATTERY:{
        'current_discharge':2,
        'capasity':3,
        'shape':4,
        'voltage':5,


        'name':1,
        'price':6,
        'url':7,
        'image':8,

        'name_table':'Батарея'
    },
    CompEnum.MICROCONTROLLER:{
        'operating_frequency': 2,
        'number_of_channels': 3,
        'operating_current': 4,
        'working_voltage': 5,
        'transmission_power':6, 
        'channel_resolution':7, 
        'wireless_protocol':8, 

        'name':1,
        'price':9,
        'url':10,
        'image':11,

        'name_table':'Микроконтроллер'
    },
    CompEnum.ELECTRICMOTOR:{
        'voltage':2,
        'maximum_power': 3,
        'recommended_battery': 4,
        'noload_current': 5,
        'peak_current': 6,
        'stator_length': 7,
        'stator_diameter':8,
        'shaft_diameter': 9,
        'number_of_revolutions_per_volt': 10,
        'resistance': 11,

        'name':1,
        'price':12,
        'url':13,
        'image':14,

        'name_table':'Электрический мотор'
    },
    CompEnum.MOTORCONTROLLER:{
        'operating_current':2,
        'peak_current':3,
        'power_support': 4,

        'name':1,
        'price':5,
        'url':6,
        'image':7,

        'name_table':'Контроллер мотора'
    },
    CompEnum.FLIGHTCONTROLLER:{
        'presence_of_a_barometer': 2,
        'presence_of_a_black_box': 3,
        'power': 4,
        'firmware': 5,
        'presence_of_a_usb_connector': 6,
        'fastening': 7,

        'name':1,
        'price':8,
        'url':9,
        'image':10,

        'name_table':'Полетный контроллер'
    },
    CompEnum.LIDAR:{
        'max_range': 2,
        'frequency': 3,
        'power_supply': 4,
        
        'name':1,
        'price':5,
        'url':6,
        'image':7,

        'name_table':'Лидар'
    },
    CompEnum.MICROFLIGHTCONTROLLER:{
        'clock_requency': 2,
        'flash_memory_capacity': 3,
        'mounting': 4,
        'min_input_voltage':5,
        'max_input_voltage': 6,
        'uart_ports_number': 7,

        'name':1,
        'price':8,
        'url':9,
        'image':10,

        'name_table':'Микроконтроллер (полетный)'
    },
    CompEnum.RANGEFINDER:{
        'max_range':2,
        'frequency': 3,
        'wave_length': 4,
        'power_supply':5,

        'name':1,
        'price':6,
        'url':7,
        'image':8,

        'name_table':'Дальномер'
    },
    CompEnum.SATELLITECOMMMODULE:{
        'battery_availability': 2,
        'battery_life': 3,
        'accuracy': 4,

        'name':1,
        'price':5,
        'url':6,
        'image':7,

        'name_table':'Модуль спутниковой связи'
        
    },
    CompEnum.LEASHINGPLATFORM:{
        'max_speed': 2,
        'gaining_speed':3,
        'deceleration_speed': 4,
        'flight_range': 5,
        'flight_altitude': 6,
        'power_consumption':7,
        'payload_weight': 8,
        'flight_time': 9,
        'screws_number': 10,

        'name':1,
        'price':11,
        'url':12,
        'image':13,

        'name_table':'Привязная платформа'
    },
    CompEnum.THERMALCAMERA:{
        'range_detection': 2,
        'range_observation': 3,
        'interface':4,
        'voltage': 5,
        'battery_availability': 6,
        'battery_life': 7,
        'field_of_view': 8,
        'magnification': 9,
        'protection_class': 10,
        'work_temperature': 11,
        'type_of_sensor': 12,

        'name':1,
        'price':13,
        'url':14,
        'image':15,

        'name_table':'Тепловизор'
    },
    CompEnum.UAVCOPTERTYPE:{
        'maximal_speed':2,
        'gaining_speed':3,
        'deceleration_speed':4,
        'maximal_range':5,
        'maximum_flight_altitude':6,
        'power_consumption':7,
        'payload_weight':8,
        'flight_time':9,
        'number_of_screws':10,

        'name':1,
        'price':11,
        'url':12,
        'image':13,

        'name_table':'БПЛА коптерного типа'
    },
    CompEnum.VIDEOTRANSMITTER:{
        'frequency': 2,
        'wattage':3,
        'number_of_channels':4,
        'antenna_connector':5,

        'name':1,
        'price':6,
        'url':7,
        'image':8,

        'name_table':'Видео передатчик'
    },
    CompEnum.PAYLOAD:{
        'matrix':2,
        'lens': 3,
        'magnification': 4,
        'number_of_megapixels': 5,
        'resolution_TVL': 6,
        'companion_image': 7,
        'thermal_imager_resolution': 8,
        'field_of_view': 9,
        'rangefinder': 10,
        'axes':11,
        'accuracy':12,
        'tangent':13,
        'roll':14,
        'yaw': 15,
        'wattage':16,
        'voltage':17,
        'current':18,
        'antenna':19,
        'frequency':20,
        'number_of_channels': 21,

        'name':1,
        'price':22,
        'url':23,
        'image':24,

        'name_table':'Полезная нагрузка'
    },
    CompEnum.CONTROLPANEL:{
        'frequency':2,
        'number_of_channels':3,
        'current':4,
        'voltage':5,
        'transmission_power':6,
        'resolution':7,
        'wireless_protocol':8,

        'name':1,
        'price':9,
        'url':10,
        'image':11,

        'name_table':'Пульт управления'
    }

}





dataset={
    CompEnum.BATTERY:{
         
            "https://fixfly.ru/dhfskdjjksdkf": 
            {
            'current_discharge':10,
            'capasity':20,
            'shape':30,
            'voltage':40,
            'url':'12345AC',
            'image':'https://fixfly.ru/images/220/20220922_163449.jpg',
            'price':3000,
            'name':'БатарейОчка'
         },
         "https://fixfly.ru/": 
            {
            'current_discharge':20,
            'capasity':20,
            'shape':30,
            'voltage':40,
            'url':'12345AC',
            'image':'https://fixfly.ru/images/220/20220922_163449.jpg',
            'price':3000,
            'name':'БатарейОчка'
         }
        
    }
}

#ws активный лист
def get_range_active_column(ws):
    column_list = [cell.column for cell in ws[1]]
    return column_list[0],len(column_list)

def parse_and_write_excel(dataset):
    global wb
    for key_component in dataset.keys():
        name=correspondence_dict[key_component]['name_table']
        ws=wb[name]
        start_index,end_index=get_range_active_column(ws)
        item=dataset[key_component]
        current_row=2
        
        for elem in item.keys():
            for key in item[elem]:
                print(item[elem])
                column_write=correspondence_dict[key_component][key]
                ws.cell(current_row, column_write).alignment=Alignment(horizontal="center")
                ws.cell(current_row, column_write).font=Font(size=11,name='Arial')
                ws.cell(current_row, column_write).value=item[elem][key]
            current_row+=1
    wb.save('result.xlsx')
    

    
            
    


if __name__ == "__main__":
    wb = load_workbook('template.xlsx')
    print(wb.sheetnames)
    parse_and_write_excel(dataset)
    # for sheet in wb.sheetnames:
    #     ws=wb[sheet]
    #     print(sheet)
    #     print("_______________________________________________")
    #     for col in ws.iter_cols(min_row=1, max_col=22, max_row=1):
    #         for cell in col:
    #             print(cell.value)
    #     # for col in ws.iter_cols(min_row=2, max_col=3, max_row=3):
    #     #     for cell in col:
    #     #         cell.value="1"
    #     #         print(cell)
    # #wb2.save('template2.xlsx')
    #         #cell.value=1
